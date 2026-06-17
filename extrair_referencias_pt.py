import os
import xml.etree.ElementTree as ET
from langdetect import detect, LangDetectException
import openpyxl

PASTA_XML = "xmls"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Referências PT por Artigo"
ws.append(["Artigo", "Nº Ref", "Título da Referência"])
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 80

for arquivo in sorted(os.listdir(PASTA_XML)):
    if not arquivo.endswith(".xml"):
        continue

    caminho = os.path.join(PASTA_XML, arquivo)

    try:
        tree = ET.parse(caminho)
        root = tree.getroot()
    except ET.ParseError:
        continue

    # extrai referências — SciELO Brasil
    refs = []
    for elem in root.findall(".//{*}citation"):
        texto = " ".join(elem.itertext()).strip()
        if len(texto) > 20:
            refs.append((elem, texto))

    # extrai referências — SciELO Portugal
    if not refs:
        for tag in ["nlm-citation", "mixed-citation", "element-citation", "ref"]:
            for elem in root.iter(tag):
                texto = " ".join(elem.itertext()).strip()
                if len(texto) > 20:
                    refs.append((elem, texto))

    refs_pt = []

    for elem, texto in refs:

        ref_lower = texto.lower()

        is_pt = False
        if any(p in ref_lower for p in [" de ", " da ", " do ", " dos ", " das ", " e ", " em "]):
            is_pt = True
        else:
            try:
                if len(texto.split()) >= 5 and detect(texto) == "pt":
                    is_pt = True
            except LangDetectException:
                pass

        if not is_pt:
            continue

        # tenta extrair título — SciELO Brasil
        titulo = ""

        # tenta article_title primeiro
        for t in elem.iter("article_title"):
            txt = " ".join(t.itertext()).strip()
            if txt:
                titulo = txt
                break

        # se não, tenta series_title
        if not titulo:
            for t in elem.iter("series_title"):
                txt = " ".join(t.itertext()).strip()
                if txt:
                    titulo = txt
                    break

        # se não, tenta CDATA (SciELO Portugal)
        if not titulo:
            for tag in ["source", "chapter-title"]:
                for t in elem.iter(tag):
                    txt = " ".join(t.itertext()).strip()
                    if txt:
                        titulo = txt
                        break

        # último recurso — texto após o ano
        if not titulo:
            partes = texto.split()
            pos_ano = -1

            for i, p in enumerate(partes):
                if p.isdigit() and len(p) == 4:
                    ano = int(p)
                    if 1900 <= ano <= 2035:
                        pos_ano = i
                        break

            if pos_ano != -1 and pos_ano + 1 < len(partes):
                titulo = " ".join(partes[pos_ano + 1:])
            else:
                titulo = texto[:150]

        if len(titulo.split()) < 5:
            continue

        refs_pt.append(titulo)

    if refs_pt:
        ws.append([arquivo, "", ""])
        ws[f"A{ws.max_row}"].font = openpyxl.styles.Font(
            bold=True,
            color="FFFFFF"
        )
        ws[f"A{ws.max_row}"].fill = openpyxl.styles.PatternFill(
            "solid",
            start_color="1F4E79"
        )

        for i, titulo in enumerate(refs_pt, 1):
            ws.append(["", i, titulo])

wb.save("referencias_pt.xlsx")
print("Arquivo 'referencias_pt.xlsx' criado!")