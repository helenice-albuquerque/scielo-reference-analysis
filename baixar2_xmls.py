import requests
import openpyxl
import os
import time
from langdetect import detect, LangDetectException

EXCEL = "Revisoes_Sistematicas_SciELO.xlsx"
PASTA_XML = "xmls"
COL_DOI = 3
TENTATIVAS = 3

os.makedirs(PASTA_XML, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL)
ws = wb.active

resultados = []

for row in ws.iter_rows(min_row=2, values_only=True):
    numero = row[0]
    titulo = row[1]
    doi_completo = row[2]

    if not doi_completo:
        continue

    doi = doi_completo.replace("https://doi.org/", "").strip()
    print(f"\n[{numero}] {doi}")

    nome_arquivo = f"{numero:02d}_{doi.replace('/', '_').replace('.', '_')}.xml"
    caminho = os.path.join(PASTA_XML, nome_arquivo)

    # pula se já foi baixado antes
    if os.path.exists(caminho):
        print(f"  → já existe, pulando")
        resultados.append((numero, titulo, doi, "✓ Já existia", nome_arquivo))
        continue

    sucesso = False
    for tentativa in range(1, TENTATIVAS + 1):
        try:
            url = f"https://articlemeta.scielo.org/api/v1/article/?code={doi}&format=xmlcrossref"
            resp = requests.get(url, timeout=20)
            if resp.status_code == 200 and len(resp.text) > 200:
                with open(caminho, "w", encoding="utf-8") as f:
                    f.write(resp.text)
                print(f"  ✓ XML baixado (tentativa {tentativa})")
                resultados.append((numero, titulo, doi, "✓ Baixado", nome_arquivo))
                sucesso = True
                break
            else:
                print(f"  ✗ Sem conteúdo (tentativa {tentativa})")
        except Exception as e:
            print(f"  ✗ Erro tentativa {tentativa}: {e}")
            time.sleep(3)

    if not sucesso:
        resultados.append((numero, titulo, doi, "✗ Falhou — verificar manualmente", ""))

    time.sleep(2)

# salva relatório
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Resultado Download"
ws2.append(["Nº", "Título", "DOI", "Status", "Arquivo XML"])
for r in resultados:
    ws2.append(r)
wb2.save("resultado_download.xlsx")

print("\n\nConcluído! Verifica o arquivo 'resultado_download.xlsx'")