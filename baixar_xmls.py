import requests
import openpyxl
import os
import time
from langdetect import detect, LangDetectException

# ── configuração ──────────────────────────────────────────────
EXCEL = "Revisoes_Sistematicas_SciELO.xlsx"  # nome do teu ficheiro
PASTA_XML = "xmls"                            # pasta onde os XMLs serão salvos
COL_DOI = 3                                   # coluna C = DOIs
# ─────────────────────────────────────────────────────────────

os.makedirs(PASTA_XML, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL)
ws = wb.active

resultados = []

for row in ws.iter_rows(min_row=2, values_only=True):
    numero = row[0]
    titulo = row[1]
    doi_completo = row[2]  # ex: https://doi.org/10.1590/...

    if not doi_completo:
        continue

    # extrai só o DOI sem o https://doi.org/
    doi = doi_completo.replace("https://doi.org/", "").strip()

    print(f"\n[{numero}] {doi}")

    # tenta API do SciELO Articlemeta
    url_api = f"https://articlemeta.scielo.org/api/v1/article/?code={doi}&format=xmlcrossref"
    
    try:
        resp = requests.get(url_api, timeout=15)
        if resp.status_code == 200 and len(resp.text) > 200:
            nome_arquivo = f"{numero:02d}_{doi.replace('/', '_').replace('.', '_')}.xml"
            caminho = os.path.join(PASTA_XML, nome_arquivo)
            with open(caminho, "w", encoding="utf-8") as f:
                f.write(resp.text)
            print(f"  ✓ XML baixado")
            resultados.append((numero, titulo, doi, "✓ Baixado", nome_arquivo))
        else:
            print(f"  ✗ API não retornou XML (status {resp.status_code})")
            resultados.append((numero, titulo, doi, f"✗ Falhou (status {resp.status_code})", ""))
    except Exception as e:
        print(f"  ✗ Erro: {e}")
        resultados.append((numero, titulo, doi, f"✗ Erro: {e}", ""))

    time.sleep(1)  # pausa para não sobrecarregar a API

# salva relatório
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Resultado Download"
ws2.append(["Nº", "Título", "DOI", "Status", "Arquivo XML"])
for r in resultados:
    ws2.append(r)
wb2.save("resultado_download.xlsx")

print("\n\nConcluído! Verifica o ficheiro 'resultado_download.xlsx'")